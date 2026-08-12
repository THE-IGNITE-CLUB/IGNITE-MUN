from flask import Blueprint, send_file, jsonify
from extensions import db
from models import Delegate, Score, Session
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm

export_bp = Blueprint('export', __name__)

def style_header(ws, headers, fill_color="131b2e"):
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

@export_bp.route('/export/delegates.xlsx', methods=['GET'])
def export_delegates_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Delegates"
    ws.row_dimensions[1].height = 20

    headers = ["#", "Name", "College", "Class", "Email", "Phone", "Committee",
               "Position 1", "Position 2", "Position 3", "Position 4", "Position 5",
               "MUN Experience", "Delegate ID", "Payment Status", "Delegation", "Registered At"]
    style_header(ws, headers)

    delegates = Delegate.query.order_by(Delegate.created_at).all()
    for i, d in enumerate(delegates, 1):
        ws.append([
            i, d.name, d.college, d.class_, d.email, d.phone, d.committee,
            d.position_1, d.position_2, d.position_3, d.position_4, d.position_5,
            d.mun_experience, d.user_id, d.payment_status, d.delegation_assigned,
            d.created_at.strftime("%Y-%m-%d %H:%M") if d.created_at else ''
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='ignite_mun_2026_delegates.xlsx')

@export_bp.route('/export/scores.xlsx', methods=['GET'])
def export_scores_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scores"
    headers = ["#", "Delegate Name", "Delegate ID", "Committee", "Session", "Decorum",
               "Policy", "Resolution", "Oratory", "Total", "Remarks"]
    style_header(ws, headers)

    scores = Score.query.all()
    for i, s in enumerate(scores, 1):
        ws.append([
            i,
            s.delegate.name if s.delegate else '',
            s.delegate.user_id if s.delegate else '',
            s.delegate.committee if s.delegate else '',
            s.session.session_type if s.session else '',
            s.decorum, s.policy, s.resolution, s.oratory, s.total, s.remarks
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='ignite_mun_2026_scores.xlsx')

@export_bp.route('/export/delegates.pdf', methods=['GET'])
def export_delegates_pdf():
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=10*mm, leftMargin=10*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#131b2e'), spaceAfter=8)
    subtitle_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#45464d'), spaceAfter=12)

    elements = [
        Paragraph("IGNITE MUN 2026 — Delegate Registry", title_style),
        Paragraph("Dialogue and Diplomacy | Sri Venkateswara University, Tirupati", subtitle_style),
        Spacer(1, 6*mm),
    ]

    delegates = Delegate.query.order_by(Delegate.created_at).all()
    data = [["#", "Name", "College", "Committee", "Delegation", "Payment", "ID"]]
    for i, d in enumerate(delegates, 1):
        data.append([str(i), d.name[:22], d.college[:20], d.committee or '', d.delegation_assigned or '-', d.payment_status, d.user_id or ''])

    t = Table(data, colWidths=[8*mm, 45*mm, 42*mm, 22*mm, 28*mm, 22*mm, 28*mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#131b2e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f2f4f6')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#c6c6cd')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)

    doc.build(elements)
    output.seek(0)
    return send_file(output, mimetype='application/pdf', as_attachment=True, download_name='ignite_mun_2026_delegates.pdf')

@export_bp.route('/export/receipt/<int:delegate_id>', methods=['GET'])
def export_payment_receipt_pdf(delegate_id):
    delegate = Delegate.query.get_or_404(delegate_id)
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()

    header_style = ParagraphStyle('RHeader', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#131b2e'), leading=22)
    sub_style = ParagraphStyle('RSub', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#006c49'), leading=14)
    normal_style = ParagraphStyle('RNorm', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#45464d'), leading=12)

    elements = [
        Paragraph("OFFICIAL PAYMENT RECEIPT", header_style),
        Paragraph("IGNITE MUN 2026 | SVU College of Engineering (SVUCE), Tirupati", sub_style),
        Spacer(1, 4*mm),
    ]

    receipt_data = [
        ["Receipt No:", f"REC-2026-{delegate.id:04d}", "Date:", delegate.created_at.strftime("%d %b %Y") if delegate.created_at else "12 Aug 2026"],
        ["Delegate ID:", delegate.user_id or "DEL-2026-001", "Payment Status:", delegate.payment_status.upper()],
        ["Delegate Name:", delegate.name, "Email:", delegate.email],
        ["Institution:", delegate.college, "Committee:", delegate.committee or "UNSC"],
        ["Transaction ID / UTR:", delegate.razorpay_payment_id or "UTR-998877665544", "Amount Paid:", "Rs. 50.00" if delegate.payment_status == 'paid' else "Rs. 0.00 (Free Slot)"],
    ]

    t = Table(receipt_data, colWidths=[35*mm, 55*mm, 35*mm, 55*mm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#191c1e')),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f2f4f6')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e3e5')),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))

    elements.append(t)
    elements.append(Spacer(1, 8*mm))
    elements.append(Paragraph("This document serves as official proof of registration payment for IGNITE MUN 2026 hosted by IGNITE Club at Sri Venkateswara University College of Engineering (SVUCE), Tirupati.", normal_style))
    elements.append(Spacer(1, 4*mm))
    elements.append(Paragraph("Authorized Stamp: [VERIFIED & STAMPED — IGNITE MUN 2026 SECRETARIAT]", ParagraphStyle('Stamp', parent=normal_style, textColor=colors.HexColor('#006c49'), fontName='Helvetica-Bold')))

    doc.build(elements)
    output.seek(0)
    return send_file(output, mimetype='application/pdf', as_attachment=True, download_name=f'IGNITE_MUN_Receipt_{delegate.user_id or "001"}.pdf')
